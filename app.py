import streamlit as st
import pandas as pd
from sheets_api import get_sheets_client_from_file, fetch_data_from_sheet
from processor import process_mis_data, create_zip_file
from matrix_engine import parse_cost_matrices, evaluate_matrix
import io
import datetime
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

SPREADSHEET_URL = os.getenv("SPREADSHEET_URL")
MIS_TAB_NAME = os.getenv("MIS_TAB_NAME", "MIS")
CONFIG_SPREADSHEET_URL = os.getenv("CONFIG_SPREADSHEET_URL")
CONFIG_TAB_NAME = os.getenv("CONFIG_TAB_NAME", "Configuration")
CUSTOM_RULES_TAB_NAME = os.getenv("CUSTOM_RULES_TAB_NAME", "Custom Rules")
COST_MATRICES_TAB_NAME = os.getenv("COST_MATRICES_TAB_NAME", "Cost Matrices")
DATE_COLUMN_NAME = os.getenv("DATE_COLUMN_NAME", "Visit Date")
BANK_COLUMN_NAME = os.getenv("BANK_COLUMN_NAME", "Bank Name")
GOOGLE_CREDENTIALS_JSON = os.getenv("GOOGLE_CREDENTIALS_JSON")
GOOGLE_CREDENTIALS_PATH = os.getenv("GOOGLE_CREDENTIALS_PATH", "service_account.json")

st.set_page_config(page_title="MIS Data Extraction", layout="wide")
st.title("MIS Billing Automation")

tab_main, tab_matrix = st.tabs(["Billing & MIS Generator", "Custom XLS Amount Calculator"])


with tab_main:
    st.sidebar.header("Data Filters")
    now = datetime.datetime.now()
    default_month = now.month - 1
    default_year = now.year

    if default_month == 0:
        default_month = 12
        default_year -= 1

    month_index = default_month - 1
    year_options = list(range(now.year - 5, now.year + 2))
    try:
        year_index = year_options.index(default_year)
    except ValueError:
        year_index = 5

    target_month = st.sidebar.selectbox("Target Month", range(1, 13), index=month_index, format_func=lambda x: datetime.date(1900, x, 1).strftime('%B'))
    target_year = st.sidebar.selectbox("Target Year", year_options, index=year_index)

    col1, col2 = st.columns(2)
    with col1:
        btn_billing = st.button("Generate Billing Sheets", use_container_width=True)
    with col2:
        btn_final = st.button("Generate Final Bills", use_container_width=True)

    if btn_billing or btn_final:
        if not SPREADSHEET_URL or SPREADSHEET_URL == "https://docs.google.com/spreadsheets/d/your_spreadsheet_id_here":
            st.error("Please set a valid SPREADSHEET_URL in your environment variables.")
        elif not CONFIG_SPREADSHEET_URL or CONFIG_SPREADSHEET_URL == "https://docs.google.com/spreadsheets/d/your_config_spreadsheet_id_here":
            st.error("Please set a valid CONFIG_SPREADSHEET_URL in your environment variables.")
        elif not GOOGLE_CREDENTIALS_JSON and not os.path.exists(GOOGLE_CREDENTIALS_PATH):
            st.error(f"Credentials not found. Please set GOOGLE_CREDENTIALS_JSON or provide the file at: '{GOOGLE_CREDENTIALS_PATH}'.")
        else:
            with st.spinner("Authenticating and Fetching Data..."):
                try:
                    # 1. Authenticate using JSON string or file path
                    if GOOGLE_CREDENTIALS_JSON:
                        from sheets_api import get_sheets_client
                        client = get_sheets_client(GOOGLE_CREDENTIALS_JSON)
                    else:
                        client = get_sheets_client_from_file(GOOGLE_CREDENTIALS_PATH)
                    
                    # 2. Fetch Data
                    mis_df = fetch_data_from_sheet(client, SPREADSHEET_URL, MIS_TAB_NAME, as_records=True)
                    
                    # Fetch Config Data (using get_all_values so it's a list of lists)
                    config_data_raw = fetch_data_from_sheet(client, CONFIG_SPREADSHEET_URL, CONFIG_TAB_NAME, as_records=False)
                    st.success("Successfully fetched data from Google Sheets.")
                    
                except Exception as e:
                    st.error(f"Error fetching data: {e}")
                    st.stop()
                    
            with st.spinner("Processing Data..."):
                try:
                    # Parse config sections
                    billing_config = []
                    final_bill_config = []
                    current_section = "BILLING"
                    
                    for row in config_data_raw:
                        if not row:
                            continue
                        first_col = str(row[0]).strip().upper()
                        if first_col == "[BILLING SHEETS]":
                            current_section = "BILLING"
                            continue
                        elif first_col == "[FINAL BILLS]":
                            current_section = "FINAL"
                            continue
                            
                        if current_section == "BILLING":
                            billing_config.append(row)
                        elif current_section == "FINAL":
                            final_bill_config.append(row)
                            
                    # Determine which config to use
                    if btn_billing:
                        target_config = billing_config
                        action_name = "Billing Sheets"
                        suffix = "_Billing"
                        zip_name = f"Billing_Sheets_{target_month}_{target_year}.zip"
                    else:
                        target_config = final_bill_config
                        action_name = "Final Bills"
                        suffix = "_Final_Bill"
                        zip_name = f"Final_Bills_{target_month}_{target_year}.zip"
                        
                    if not target_config:
                        st.warning(f"No configuration found for {action_name}. Ensure you have added rows under the correct marker.")
                    else:
                        # 3. Process Data
                        custom_rules_df = pd.DataFrame()
                        try:
                            custom_rules_df = fetch_data_from_sheet(client, CONFIG_SPREADSHEET_URL, CUSTOM_RULES_TAB_NAME, as_records=True)
                        except Exception as e:
                            st.warning(f"Could not load the '{CUSTOM_RULES_TAB_NAME}' tab from Configuration sheet. Ensure the name is exactly matching. ({e})")
                            
                        files, logs = process_mis_data(mis_df, target_config, target_month, target_year, DATE_COLUMN_NAME, BANK_COLUMN_NAME, custom_rules_df)
                        
                        # Show execution logs
                        st.subheader(f"Execution Log: {action_name}")
                        with st.container(border=True):
                            for log in logs:
                                st.text(log)
                                
                        if not files:
                            st.warning("No data found for the given month/year or matching bank configurations.")
                        else:
                            # 4. Create ZIP
                            zip_bytes = create_zip_file(files, suffix=suffix)
                            
                            st.success(f"Successfully generated files for {len(files)} banks!")
                            st.download_button(
                                label=f"⬇️ Download {action_name} as ZIP",
                                data=zip_bytes,
                                file_name=zip_name,
                                mime="application/zip",
                                type="primary"
                            )
                                
                except Exception as e:
                    st.error(f"Error processing data: {e}")


with tab_matrix:
    st.header("Calculate Amounts for Custom XLS")
    st.write("Upload an Excel file to automatically calculate and populate the Amount column based on the Cost Matrices defined in Google Sheets.")
    
    if not CONFIG_SPREADSHEET_URL or not GOOGLE_CREDENTIALS_JSON and not os.path.exists(GOOGLE_CREDENTIALS_PATH):
        st.warning("Please configure your Google Sheets and Credentials to use this feature.")
    else:
        # Load matrices
        matrices = {}
        try:
            if GOOGLE_CREDENTIALS_JSON:
                from sheets_api import get_sheets_client
                client = get_sheets_client(GOOGLE_CREDENTIALS_JSON)
            else:
                client = get_sheets_client_from_file(GOOGLE_CREDENTIALS_PATH)
            
            raw_matrix_data = fetch_data_from_sheet(client, CONFIG_SPREADSHEET_URL, COST_MATRICES_TAB_NAME, as_records=False)
            matrices = parse_cost_matrices(raw_matrix_data)
        except Exception as e:
            st.error(f"Error loading Cost Matrices tab: {e}")
            
        if not matrices:
            st.info("No matrices found. Please configure the 'Cost Matrices' tab.")
        else:
            selected_bank = st.selectbox("Select Bank Configuration", list(matrices.keys()))
            uploaded_file = st.file_uploader("Upload Excel File (.xls or .xlsx)", type=['xls', 'xlsx'])
            
            if uploaded_file is not None:
                try:
                    xls = pd.ExcelFile(uploaded_file)
                    sheet_names = xls.sheet_names
                except Exception as e:
                    st.error(f"Error reading Excel file: {e}")
                    sheet_names = []
                    
                if sheet_names:
                    selected_sheet = st.selectbox("Select Sheet to Process", sheet_names)
                    
                    if st.button("Calculate & Download"):
                        with st.spinner("Processing file..."):
                            try:
                                df = pd.read_excel(uploaded_file, sheet_name=selected_sheet)
                                matrix_def = matrices[selected_bank]
                                
                                df_processed, matrix_logs = evaluate_matrix(df, matrix_def)
                                
                                st.subheader("Execution Log")
                                with st.container(border=True):
                                    for log in matrix_logs:
                                        st.text(log)
                                        
                                # Save to memory and preserve formatting using openpyxl
                                import openpyxl
                                uploaded_file.seek(0)
                                wb = openpyxl.load_workbook(uploaded_file)
                                ws = wb[selected_sheet]
                                
                                amount_col_name = matrix_def.get('output_col', 'Amount')
                                
                                # 1. Find the header row by looking for the first column of our dataframe
                                header_row = 1
                                if len(df.columns) > 0:
                                    first_col_name = str(df.columns[0]).strip()
                                    found = False
                                    for r in range(1, 21): # Search first 20 rows
                                        for c in range(1, ws.max_column + 1):
                                            val = ws.cell(row=r, column=c).value
                                            if val is not None and str(val).strip() == first_col_name:
                                                header_row = r
                                                found = True
                                                break
                                        if found:
                                            break
                                            
                                # 2. Find the target amount column index
                                amount_col_idx = None
                                for c in range(1, ws.max_column + 2):
                                    val = ws.cell(row=header_row, column=c).value
                                    if val is not None and str(val).strip() == amount_col_name:
                                        amount_col_idx = c
                                        break
                                        
                                if not amount_col_idx:
                                    amount_col_idx = ws.max_column + 1
                                    ws.cell(row=header_row, column=amount_col_idx).value = amount_col_name
                                    
                                # 3. Write values sequentially
                                for i, val in enumerate(df_processed[amount_col_name]):
                                    ws.cell(row=header_row + 1 + i, column=amount_col_idx).value = val
                                    
                                output = io.BytesIO()
                                wb.save(output)
                                output.seek(0)
                                
                                original_name = uploaded_file.name
                                new_name = original_name.rsplit('.', 1)[0] + "_Priced.xlsx"
                                
                                st.success("File processed successfully!")
                                st.download_button(
                                    label="⬇️ Download Processed Excel",
                                    data=output.getvalue(),
                                    file_name=new_name,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    type="primary"
                                )
                            except Exception as e:
                                st.error(f"Error processing uploaded file: {e}")
